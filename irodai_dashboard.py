import streamlit as strl
import firebase_admin
from firebase_admin import credentials, firestore
import pandas as pd
from datetime import datetime, time, timedelta
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Oldal konfigurációja (szélesvásznú asztali nézet)
strl.set_page_config(
    page_title="FRD Raktár - Vezetői Dashboard",
    page_icon="📊",
    layout="wide"
)

# 1. Firebase csatlakozás inicializálása
@strl.cache_resource
def init_firebase():
    try:
        if "p_key" not in strl.secrets:
            return None
            
        key_dict = {
            "type": "service_account",
            "project_id": "frd-alapanyag",
            "private_key_id": strl.secrets["p_id"],
            "private_key": strl.secrets["p_key"],
            "client_email": "firebase-adminsdk-fbsvc@frd-alapanyag.iam.gserviceaccount.com",
            "client_id": "118377480036110848051",
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
            "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-fbsvc%40frd-alapanyag.iam.gserviceaccount.com",
            "universe_domain": "googleapis.com"
        }
            
        cred = credentials.Certificate(key_dict)
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
            
        return firestore.client()
    except Exception as e:
        strl.error(f"X Rendszerhiba a kapcsolódáskor: {e}")
        return None

db = init_firebase()

# 2. Adatok letöltése a Firestore-ból
@strl.cache_data(ttl=10)
def get_raktar_adatok():
    if db is None:
        return []
    try:
        docs_snapshot = db.collection('materials').get()
        adatok = []
        for doc in docs_snapshot:
            d = doc.to_dict()
            try:
                current = float(d.get('currentStock', 0))
                minimum = float(d.get('minStock', 20))
            except:
                current, minimum = 0.0, 20.0

            adatok.append({
                "Cikkszám (SKU)": d.get('sku', doc.id),
                "Megnevezés": d.get('name', 'Névtelen alapanyag'),
                "Kategória": d.get('type', 'Egyéb'),
                "Készlet": int(current) if current % 1 == 0 else current,
                "Minimum szint": int(minimum) if minimum % 1 == 0 else minimum,
                "Egység": d.get('unit', 'Pár'),
                "Státusz": "🚨 HIÁNY" if current <= minimum else "✅ Rendben"
            })
        return adatok
    except Exception as e:
        strl.error(f"X HIBA az adatok letöltése közben: {e}")
        return []

# 3. Raktári naplófájlok (logs) lekérése – Dátumintervallum alapú szűréssel
@strl.cache_data(ttl=5)
def get_idoszakos_mozgasok(start_date, end_date):
    if db is None:
        return []
    try:
        start_datetime = datetime.combine(start_date, time.min)
        end_datetime = datetime.combine(end_date, time.max)
        
        logs_snapshot = db.collection('logs')\
            .where('timestamp', '>=', start_datetime)\
            .where('timestamp', '<=', end_datetime)\
            .get()
            
        mozgasok = []
        tipus_leforditott = {
            "felhasznalas": "Felhasználás",
            "selejt": "Selejt",
            "atgolyositas_ki": "Átalakítás (Kiadás)",
            "atgolyositas_be": "Átalakítás (Beérkezés)"
        }
        
        for doc in logs_snapshot:
            d = doc.to_dict()
            nyers_tipus = d.get('logType', 'felhasznalas')
            mennyiseg = float(d.get('quantity', 0))
            
            mozgasok.append({
                "Cikkszám (SKU)": d.get('sku', '-'),
                "Megnevezés": d.get('name', 'Névtelen'),
                "Kategória": d.get('type', 'Egyéb'),
                "Típus": tipus_leforditott.get(nyers_tipus, nyers_tipus),
                "Mennyiség": int(mennyiseg) if mennyiseg % 1 == 0 else mennyiseg,
                "Egység": d.get('unit', 'pár')
            })
            
        return mozgasok
    except Exception as e:
        strl.error(f"X HIBA a naplózott adatok letöltése közben: {e}")
        return []

# 4. Kifinomult Excel generáló – Időszaki összesítéssel és intelligens formázással
def generalk_idoszakos_excel(df_keszlet, df_mozgasok, idoszak_str):
    excel_buffer = BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_keszlet.to_excel(writer, index=False, sheet_name='Aktuális Készlet')
        
        workbook = writer.book
        sheet_name_mozgas = 'Időszaki Mozgások'
        ws2 = workbook.create_sheet(title=sheet_name_mozgas)
        ws2.views.sheetView[0].showGridLines = True
        
        # Stílusok
        font_main_title = Font(name="Calibri", size=14, bold=True, color="1F4E78")
        font_section_title = Font(name="Calibri", size=12, bold=True, color="000000")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_italic_info = Font(name="Calibri", size=11, italic=True, color="595959")
        
        fill_keszlet_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_felhasznalas = PatternFill(start_color="2E5B82", end_color="2E5B82", fill_type="solid")
        fill_selejt = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        fill_atalakitas = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        
        # --- 1. FÜL FORMÁZÁSA ---
        ws1 = workbook['Aktuális Készlet']
        ws1.row_dimensions[1].height = 26
        for cell in ws1[1]:
            cell.fill = fill_keszlet_header
            cell.font = font_header
            cell.alignment = align_center
            
        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
            if col_letter in ['A', 'D', 'E', 'F', 'G']:
                for cell in col[1:]: cell.alignment = Alignment(horizontal="center")
            else:
                for cell in col[1:]: cell.alignment = Alignment(horizontal="left")

        # --- 2. FÜL FORMÁZÁSA ---
        ws2['A1'] = f"IDŐSZAKOS RAKTÁRMOZGÁSI ÖSSZESÍTŐ ({idoszak_str})"
        ws2['A1'].font = font_main_title
        ws2.row_dimensions[1].height = 25
        
        current_row = 3
        
        def beszuro_osszesitett_tabla(szurt_df, szekcio_nev, fejlec_fill):
            nonlocal current_row
            
            ws2.cell(row=current_row, column=1, value=szekcio_nev).font = font_section_title
            ws2.row_dimensions[current_row].height = 20
            current_row += 1
            
            headers = ["Cikkszám (SKU)", "Megnevezés", "Kategória", "Típus", "Összesített Mennyiség", "Egység"]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws2.cell(row=current_row, column=col_idx, value=header)
                cell.font = font_header
                cell.fill = fejlec_fill
                cell.alignment = align_center
            ws2.row_dimensions[current_row].height = 24
            current_row += 1
            
            if szurt_df.empty:
                cell = ws2.cell(row=current_row, column=1, value="A megadott időszakban nem történt ilyen mozgás.")
                cell.font = font_italic_info
                ws2.row_dimensions[current_row].height = 18
                current_row += 3
                return
            
            # ÖSSZESÍTÉS LOGIKÁJA: Cikkszám szerint összeadjuk a mennyiségeket az időszakra
            grouped = szurt_df.groupby(["Cikkszám (SKU)", "Megnevezés", "Kategória", "Típus", "Egység"])["Mennyiség"].sum().reset_index()
            # Sorrend: kategória és mennyiség szerint csökkenőben (hogy a legtöbbet használt legyen elöl)
            grouped = grouped.sort_values(by=["Kategória", "Mennyiség"], ascending=[True, False])
            
            for _, row_data in grouped.iterrows():
                row_values = [
                    row_data["Cikkszám (SKU)"],
                    row_data["Megnevezés"],
                    row_data["Kategória"],
                    row_data["Típus"],
                    row_data["Mennyiség"],
                    row_data["Egység"]
                ]
                for col_idx, val in enumerate(row_values, 1):
                    cell = ws2.cell(row=current_row, column=col_idx, value=val)
                    header_name = headers[col_idx - 1]
                    if header_name in ["Cikkszám (SKU)", "Típus", "Összesített Mennyiség", "Egység"]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                ws2.row_dimensions[current_row].height = 18
                current_row += 1
                
            current_row += 2

        if not df_mozgasok.empty:
            df_felhasznalas = df_mozgasok[df_mozgasok["Típus"] == "Felhasználás"]
            df_selejt = df_mozgasok[df_mozgasok["Típus"] == "Selejt"]
            df_atalakitas = df_mozgasok[df_mozgasok["Típus"].str.contains("Átalakítás")]
        else:
            df_felhasznalas = df_selejt = df_atalakitas = pd.DataFrame()

        beszuro_osszesitett_tabla(df_felhasznalas, "1. Időszaki Összesített Felhasználások (Legtöbbet fogyott tételek elöl)", fill_felhasznalas)
        beszuro_osszesitett_tabla(df_selejt, "2. Időszaki Összesített Selejtek (Kritikus veszteségek követése)", fill_selejt)
        beszuro_osszesitett_tabla(df_atalakitas, "3. Időszaki Összesített Átalakítások", fill_atalakitas)

        for col in ws2.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value and cell.coordinate != 'A1':
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

    return excel_buffer.getvalue()

# --- UI FELÉPÍTÉSE ---
strl.title("📊 FRD Alapanyag Raktár - Vezetői Műszerfal")
strl.caption("Élő, irodai betekintő felület az üzemben lévő tabletek készletéhez és időszaki jelentésekhez")
strl.write("---")

if db is not None:
    nyers_adatok = get_raktar_adatok()
    df = pd.DataFrame(nyers_adatok)

    if not df.empty:
        hianyzo_df = df[df["Státusz"] == "🚨 HIÁNY"]
        
        col1, col2, col3 = strl.columns([1, 1, 1])
        with col1:
            strl.metric(label="Összes egyedi alapanyag", value=len(df))
        with col2:
            strl.metric(
                label="Készlethiányos tételek száma", 
                value=len(hianyzo_df),
                delta=f"{len(hianyzo_df)} azonnali beszerzés" if len(hianyzo_df) > 0 else "Minden rendben",
                delta_color="inverse" if len(hianyzo_df) > 0 else "normal"
            )
        with col3:
            # ÚJ: Időszak típus kiválasztása a menedzsmentnek
            ma = datetime.now().date()
            idoszak_tipus = strl.selectbox(
                "Válaszd ki az elemzési időszakot:",
                ["Egy konkrét nap", "Utolsó 30 nap (1 hónap)", "Utolsó 90 nap (Negyedév)", "Utolsó 180 nap (Fél év)"]
            )
            
            # Dátumok kiszámítása a háttérben
            if idoszak_tipus == "Egy konkrét nap":
                valasztott_datum = strl.date_input("Válaszd ki a napot:", ma)
                start_date = end_date = valasztott_datum
                idoszak_str = start_date.strftime('%Y-%m-%d')
            elif idoszak_tipus == "Utolsó 30 nap (1 hónap)":
                start_date = ma - timedelta(days=30)
                end_date = ma
                idoszak_str = f"{start_date.strftime('%Y-%m-%d')} - től {end_date.strftime('%Y-%m-%d')}-ig"
            elif idoszak_tipus == "Utolsó 90 nap (Negyedév)":
                start_date = ma - timedelta(days=90)
                end_date = ma
                idoszak_str = f"{start_date.strftime('%Y-%m-%d')} - től {end_date.strftime('%Y-%m-%d')}-ig"
            else:  # Fél év
                start_date = ma - timedelta(days=180)
                end_date = ma
                idoszak_str = f"{start_date.strftime('%Y-%m-%d')} - től {end_date.strftime('%Y-%m-%d')}-ig"

        strl.write("---")
        
        # Lekérjük az adatokat a kalkulált intervallumra
        nyers_mozgasok = get_idoszakos_mozgasok(start_date, end_date)
        df_mozgasok = pd.DataFrame(nyers_mozgasok)

        strl.subheader(f"📈 Kiválasztott időszak adatai: {idoszak_tipus}")
        strl.info(f"Aktív szűrési tartomány: **{idoszak_str}**")
        
        rep_col1, rep_col2 = strl.columns([2, 1])
        with rep_col1:
            if not df_mozgasok.empty:
                strl.write(f"A kijelölt időszakban összesen **{len(df_mozgasok)} nyers tranzakció** történt az üzemben. Az Excel letöltéskor ezeket automatikusan alkatrészenként összesítve kapod meg!")
            else:
                strl.info("A kijelölt időintervallumban nem található rögzített mozgás az adatbázisban.")
        
        with rep_col2:
            excel_adatok = generalk_idoszakos_excel(df, df_mozgasok, idoszak_str)
            strl.download_button(
                label=f"📥 Időszaki Összesített Excel Letöltése",
                data=excel_adatok,
                file_name=f"frd_raktar_osszesito_{ma.strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        # Képernyős betekintő (Ha hosszabb időszak van kijelölve, itt is összesítve mutatjuk meg, hogy átlátható legyen)
        if not df_mozgasok.empty:
            with strl.expander("👀 Összesített adatok gyors megtekintése a képernyőn", expanded=True):
                kepernyos_szumma = df_mozgasok.groupby(["Cikkszám (SKU)", "Megnevezés", "Kategória", "Típus", "Egység"])["Mennyiség"].sum().reset_index()
                kepernyos_szumma = kepernyos_szumma.sort_values(by=["Típus", "Kategória", "Mennyiség"], ascending=[True, True, False])
                
                mozgas_szuro = strl.multiselect("Szűrés típus szerint a képernyőn:", list(kepernyos_szumma["Típus"].unique()), default=list(kepernyos_szumma["Típus"].unique()))
                megjelenitendo_mozgas_df = kepernyos_szumma[kepernyos_szumma["Típus"].isin(mozgas_szuro)]
                
                # Átnevezzük az oszlopot a képernyőn is a tisztaság kedvéért
                megjelenitendo_mozgas_df = megjelenitendo_mozgas_df.rename(columns={"Mennyiség": "Időszaki Összes Mennyiség"})
                strl.dataframe(megjelenitendo_mozgas_df, use_container_width=True, hide_index=True)

        strl.write("---")

        if not hianyzo_df.empty:
            strl.error("### 🚨 Az alábbi alapanyagok készlete a kritikus minimum alá süllyedt!")
            strl.dataframe(hianyzo_df, use_container_width=True, hide_index=True)
            strl.write("---")

        strl.subheader("🔍 Keresés és szűrés a teljes aktuális raktárban")
        f_col1, f_col2 = strl.columns([1, 2])
        
        with f_col1:
            kategoriak = ["Mind"] + sorted(list(df["Kategória"].unique()))
            valasztott_kat = strl.selectbox("Szűrés kategória szerint:", kategoriak)
            
        with f_col2:
            kereses = strl.text_input("Keresés név vagy cikkszám alapján:", "").strip().lower()

        megjelenitendo_df = df.copy()
        if valasztott_kat != "Mind":
            megjelenitendo_df = megjelenitendo_df[megjelenitendo_df["Kategória"] == valasztott_kat]
        if kereses:
            megjelenitendo_df = megjelenitendo_df[
                megjelenitendo_df["Megnevezés"].str.lower().str.contains(kereses) | 
                megjelenitendo_df["Cikkszám (SKU)"].str.lower().str.contains(kereses)
            ]

        strl.dataframe(megjelenitendo_df, use_container_width=True, hide_index=True)
